from openpyxl import load_workbook
from random import shuffle

wb = load_workbook("all_lyrics.xlsx")
ws = wb.active

dataset = list(zip(ws['B'], ws['C']))           # создаем датасет из пар title и lyrics
shuffle(dataset)                                # и перемешиваем чтобы разделить на группы
size = len(dataset)

train_dataset = dataset[0: int(0.9 * size)]     # 90% песен идут в тренировочный
val_dataset = dataset[int(0.9 * size): ]        # 10% в валидационный

def helper(dataset, type):
    """
    dataset: лист из пар title, lyrics для тренировки или валидации 
    type: равна либо "train" либо "var"; чисто для названия файла
    """
    dataset_text = ""
    for title, lyrics in dataset:
        object = (f"<s>Название песни: {title.value}"
                  f"\nТекст песни:\n{lyrics.value}</s>\n")
        dataset_text += object
    with open(type + ".txt", 'w', encoding="utf-8") as f:
        f.write(dataset_text)

# создаем файлы train.txt и val.txt для тренировки ruGPT3
for dataset, type in [(train_dataset, "train"), (val_dataset, "val")]:
    helper(dataset=dataset, type=type)