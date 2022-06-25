from openpyxl import load_workbook
from random import shuffle
import yaml

config = config = yaml.safe_load(open('config/params.yaml'))
train_size      = config['train']['train_size']
dataset_file    = config['path']['dataset']
train_path      = config['path']['train']
val_path        = config['path']['val']

def helper(dataset, path):
    """
    dataset: лист из пар title, lyrics для тренировки или валидации 
    type: равна либо "train" либо "var"; чисто для названия файла
    """
    dataset_text = ""
    for title, lyrics in dataset:
        object = (f"<s>Название песни: {title.value}"
                  f"\nТекст песни:\n{lyrics.value}</s>\n")
        dataset_text += object
    with open(path, 'w', encoding="utf-8") as f:
        f.write(dataset_text)

wb = load_workbook(dataset_file)
ws = wb.active

dataset = list(zip(ws['B'], ws['C']))           # создаем датасет из пар title и lyrics
shuffle(dataset)                                # и перемешиваем чтобы разделить на группы
size = len(dataset)

train_dataset = dataset[0: int(train_size * size)]     # train_size % песен идут в тренировочный
val_dataset = dataset[int(train_size * size): ]        # (1 - train_size) % в валидационный

# создаем файлы train.txt и val.txt для тренировки ruGPT3
for dataset, path in [(train_dataset, train_path), (val_dataset, val_path)]:
    helper(dataset=dataset, path=path)